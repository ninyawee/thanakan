"""Export functions for account data (JSON, CSV, Excel)."""

import csv
import json
from pathlib import Path

import openpyxl

from .models import Account


def export_to_json(accounts: list[Account], output_path: Path | str) -> None:
    """Export accounts to JSON file.

    Args:
        accounts: List of Account objects
        output_path: Path to output JSON file
    """
    output_path = Path(output_path)
    data = [acc.model_dump(mode="json") for acc in accounts]
    with open(output_path, "w") as f:
        json.dump(data, f, indent=2, default=str)


def export_to_csv(accounts: list[Account], output_dir: Path | str) -> None:
    """Export accounts to CSV files (one per account).

    Args:
        accounts: List of Account objects
        output_dir: Directory for output CSV files
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    for account in accounts:
        filename = f"{account.account_number.replace('-', '')}.csv"
        output_path = output_dir / filename

        with open(output_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)

            # Header
            writer.writerow([
                "Date",
                "Time",
                "Description",
                "Channel",
                "Check Number",
                "Withdrawal",
                "Deposit",
                "Balance",
                "Reference",
            ])

            # Transactions
            for txn in account.all_transactions:
                writer.writerow([
                    str(txn.date),
                    str(txn.time) if txn.time else "",
                    txn.description,
                    txn.channel or "",
                    txn.check_number or "",
                    float(txn.withdrawal) if txn.withdrawal else "",
                    float(txn.deposit) if txn.deposit else "",
                    float(txn.balance),
                    txn.reference or "",
                ])


def export_to_excel(accounts: list[Account], output_path: Path | str) -> None:
    """Export accounts to Excel file.

    Each account gets its own sheet.

    Args:
        accounts: List of Account objects
        output_path: Path to output Excel file
    """
    output_path = Path(output_path)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    for account in accounts:
        # Create sheet with account number (sanitized for Excel)
        sheet_name = account.account_number.replace("-", "")[:31]
        ws = wb.create_sheet(sheet_name)

        # Header
        headers = [
            "Date",
            "Time",
            "Description",
            "Channel",
            "Withdrawal",
            "Deposit",
            "Balance",
            "Reference",
        ]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # Transactions
        for row, txn in enumerate(account.all_transactions, 2):
            ws.cell(row=row, column=1, value=str(txn.date))
            ws.cell(row=row, column=2, value=str(txn.time) if txn.time else "")
            ws.cell(row=row, column=3, value=txn.description)
            ws.cell(row=row, column=4, value=txn.channel or "")
            ws.cell(row=row, column=5, value=float(txn.withdrawal) if txn.withdrawal else None)
            ws.cell(row=row, column=6, value=float(txn.deposit) if txn.deposit else None)
            ws.cell(row=row, column=7, value=float(txn.balance))
            ws.cell(row=row, column=8, value=txn.reference or "")

    wb.save(output_path)
