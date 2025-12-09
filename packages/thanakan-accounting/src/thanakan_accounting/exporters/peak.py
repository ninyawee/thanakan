"""Peak Import Statement exporter.

Exports bank statement data to Peak accounting software format.
Peak Import Statement uses an Excel file with 3 columns:
- Column A: Date (YYYYMMDD format)
- Column B: Amount (negative for withdrawals, positive for deposits)
- Column C: Note (description, max 1000 characters)
"""

from pathlib import Path

import openpyxl

from thanakan_statement import Account, Transaction

# Peak Import Statement template format
SHEET_NAME = "Import_BankStatement"
HEADERS = ("วันที่รายการ *", "จำนวนเงิน *", "หมายเหตุ")


def _format_date(txn: Transaction) -> str:
    """Format transaction date to YYYYMMDD format for Peak."""
    return txn.date.strftime("%Y%m%d")


def _format_amount(txn: Transaction) -> float:
    """Format transaction amount (negative for withdrawal, positive for deposit)."""
    if txn.withdrawal is not None:
        return -float(txn.withdrawal)
    elif txn.deposit is not None:
        return float(txn.deposit)
    return 0.0


def _format_note(txn: Transaction, max_length: int = 1000) -> str:
    """Format transaction note from description, channel, check_number, and reference.

    Args:
        txn: Transaction object
        max_length: Maximum note length (Peak allows 1000 characters)

    Returns:
        Combined note string
    """
    parts = [txn.description]
    if txn.channel:
        parts.append(f"[{txn.channel}]")
    if txn.check_number:
        parts.append(f"Chq#{txn.check_number}")
    if txn.reference:
        parts.append(txn.reference)

    note = " ".join(parts)
    return note[:max_length]


def _write_transactions(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    transactions: list[Transaction],
) -> None:
    """Write transactions to worksheet with header row."""
    # Write header at row 1
    for col, header in enumerate(HEADERS, 1):
        ws.cell(row=1, column=col, value=header)

    # Data starts at row 2
    for row_num, txn in enumerate(transactions, 2):
        ws.cell(row=row_num, column=1, value=_format_date(txn))
        ws.cell(row=row_num, column=2, value=_format_amount(txn))
        ws.cell(row=row_num, column=3, value=_format_note(txn))

    # Adjust column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 80


def export_to_peak(accounts: list[Account], output_path: Path | str) -> None:
    """Export multiple accounts to Peak Import Statement Excel format.

    Each account is exported to a separate sheet.

    Args:
        accounts: List of Account objects with consolidated transactions
        output_path: Path to output Excel file
    """
    output_path = Path(output_path)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    for account in accounts:
        sheet_name = account.account_number.replace("-", "")[:31]
        ws = wb.create_sheet(sheet_name)
        _write_transactions(ws, account.all_transactions)

    wb.save(output_path)


def export_single_to_peak(account: Account, output_path: Path | str) -> None:
    """Export a single account to Peak Import Statement format.

    Args:
        account: Account object with consolidated transactions
        output_path: Path to output Excel file
    """
    output_path = Path(output_path)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    _write_transactions(ws, account.all_transactions)
    wb.save(output_path)
