"""Tests for Peak Import Statement exporter."""

import datetime as dt
from decimal import Decimal

import openpyxl
import pytest

from thanakan_accounting.exporters.peak import (
    HEADERS,
    SHEET_NAME,
    _format_amount,
    _format_date,
    _format_note,
    export_single_to_peak,
    export_to_peak,
)
from thanakan_statement import Account, Transaction


@pytest.fixture
def sample_transaction() -> Transaction:
    """Create a sample withdrawal transaction."""
    return Transaction(
        date=dt.date(2025, 11, 1),
        time=dt.time(10, 30),
        description="Transfer Withdrawal",
        channel="K PLUS",
        withdrawal=Decimal("8400.00"),
        deposit=None,
        balance=Decimal("50000.00"),
        reference="REF123456",
        check_number=None,
    )


@pytest.fixture
def sample_deposit_transaction() -> Transaction:
    """Create a sample deposit transaction."""
    return Transaction(
        date=dt.date(2025, 11, 2),
        time=dt.time(14, 0),
        description="Transfer Deposit",
        channel="K PLUS",
        withdrawal=None,
        deposit=Decimal("15000.50"),
        balance=Decimal("65000.50"),
        reference=None,
        check_number=None,
    )


@pytest.fixture
def sample_check_transaction() -> Transaction:
    """Create a sample check transaction."""
    return Transaction(
        date=dt.date(2025, 11, 3),
        time=None,
        description="Check Payment",
        channel=None,
        withdrawal=Decimal("5000.00"),
        deposit=None,
        balance=Decimal("60000.50"),
        reference=None,
        check_number="12345678",
    )


@pytest.fixture
def sample_account(sample_transaction, sample_deposit_transaction) -> Account:
    """Create a sample account with transactions."""
    return Account(
        account_number="123-4-56789-0",
        account_name="Test Account",
        all_transactions=[sample_transaction, sample_deposit_transaction],
    )


class TestFormatDate:
    """Tests for _format_date function."""

    def test_format_date_basic(self, sample_transaction):
        """Date should be formatted as YYYYMMDD."""
        result = _format_date(sample_transaction)
        assert result == "20251101"

    def test_format_date_december(self):
        """December dates should format correctly."""
        txn = Transaction(
            date=dt.date(2025, 12, 31),
            description="Test",
            balance=Decimal("1000"),
        )
        result = _format_date(txn)
        assert result == "20251231"

    def test_format_date_january(self):
        """January dates with leading zeros should format correctly."""
        txn = Transaction(
            date=dt.date(2025, 1, 5),
            description="Test",
            balance=Decimal("1000"),
        )
        result = _format_date(txn)
        assert result == "20250105"


class TestFormatAmount:
    """Tests for _format_amount function."""

    def test_withdrawal_is_negative(self, sample_transaction):
        """Withdrawals should be negative amounts."""
        result = _format_amount(sample_transaction)
        assert result == -8400.0

    def test_deposit_is_positive(self, sample_deposit_transaction):
        """Deposits should be positive amounts."""
        result = _format_amount(sample_deposit_transaction)
        assert result == 15000.50

    def test_zero_when_neither(self):
        """Return 0 when neither withdrawal nor deposit."""
        txn = Transaction(
            date=dt.date(2025, 11, 1),
            description="Interest",
            withdrawal=None,
            deposit=None,
            balance=Decimal("1000"),
        )
        result = _format_amount(txn)
        assert result == 0.0


class TestFormatNote:
    """Tests for _format_note function."""

    def test_description_only(self):
        """Note with description only."""
        txn = Transaction(
            date=dt.date(2025, 11, 1),
            description="Simple Transfer",
            balance=Decimal("1000"),
        )
        result = _format_note(txn)
        assert result == "Simple Transfer"

    def test_with_channel(self, sample_transaction):
        """Note with description and channel."""
        result = _format_note(sample_transaction)
        assert "Transfer Withdrawal" in result
        assert "[K PLUS]" in result
        assert "REF123456" in result

    def test_with_check_number(self, sample_check_transaction):
        """Note with check number."""
        result = _format_note(sample_check_transaction)
        assert "Check Payment" in result
        assert "Chq#12345678" in result

    def test_max_length_truncation(self, sample_transaction):
        """Note should be truncated to max_length."""
        result = _format_note(sample_transaction, max_length=20)
        assert len(result) <= 20

    def test_all_fields_combined(self):
        """Note with all fields present."""
        txn = Transaction(
            date=dt.date(2025, 11, 1),
            description="Transfer",
            channel="ATM",
            withdrawal=Decimal("1000"),
            balance=Decimal("5000"),
            reference="REF999",
            check_number="CHK001",
        )
        result = _format_note(txn)
        assert "Transfer" in result
        assert "[ATM]" in result
        assert "Chq#CHK001" in result
        assert "REF999" in result


class TestExportSingleToPeak:
    """Tests for export_single_to_peak function."""

    def test_creates_excel_file(self, sample_account, tmp_path):
        """Should create an Excel file."""
        output = tmp_path / "test_output.xlsx"
        export_single_to_peak(sample_account, output)
        assert output.exists()

    def test_correct_sheet_name(self, sample_account, tmp_path):
        """Sheet name should be Import_BankStatement."""
        output = tmp_path / "test_output.xlsx"
        export_single_to_peak(sample_account, output)

        wb = openpyxl.load_workbook(output)
        assert wb.active.title == SHEET_NAME
        wb.close()

    def test_header_row(self, sample_account, tmp_path):
        """First row should contain Thai headers."""
        output = tmp_path / "test_output.xlsx"
        export_single_to_peak(sample_account, output)

        wb = openpyxl.load_workbook(output)
        ws = wb.active
        header = (ws.cell(1, 1).value, ws.cell(1, 2).value, ws.cell(1, 3).value)
        assert header == HEADERS
        wb.close()

    def test_data_starts_row_2(self, sample_account, tmp_path):
        """Data should start at row 2."""
        output = tmp_path / "test_output.xlsx"
        export_single_to_peak(sample_account, output)

        wb = openpyxl.load_workbook(output)
        ws = wb.active
        # Row 2 should have transaction data
        assert ws.cell(2, 1).value == "20251101"  # Date
        assert ws.cell(2, 2).value == -8400.0  # Amount (withdrawal)
        wb.close()

    def test_transaction_count(self, sample_account, tmp_path):
        """Should have correct number of transaction rows."""
        output = tmp_path / "test_output.xlsx"
        export_single_to_peak(sample_account, output)

        wb = openpyxl.load_workbook(output)
        ws = wb.active
        # Header + 2 transactions = 3 rows with data
        assert ws.cell(3, 1).value == "20251102"  # Second transaction
        assert ws.cell(4, 1).value is None  # No more data
        wb.close()


class TestExportToPeak:
    """Tests for export_to_peak function (multi-account)."""

    def test_creates_file(self, sample_account, tmp_path):
        """Should create an Excel file."""
        output = tmp_path / "multi_account.xlsx"
        export_to_peak([sample_account], output)
        assert output.exists()

    def test_multiple_accounts_multiple_sheets(
        self, sample_transaction, sample_deposit_transaction, tmp_path
    ):
        """Each account should get its own sheet."""
        account1 = Account(
            account_number="111-1-11111-1",
            all_transactions=[sample_transaction],
        )
        account2 = Account(
            account_number="222-2-22222-2",
            all_transactions=[sample_deposit_transaction],
        )

        output = tmp_path / "multi_account.xlsx"
        export_to_peak([account1, account2], output)

        wb = openpyxl.load_workbook(output)
        assert len(wb.sheetnames) == 2
        assert "1111111111" in wb.sheetnames  # Dashes removed
        assert "2222222222" in wb.sheetnames
        wb.close()

    def test_sheet_name_truncation(self, sample_transaction, tmp_path):
        """Sheet names longer than 31 chars should be truncated."""
        account = Account(
            account_number="123456789012345678901234567890123456",
            all_transactions=[sample_transaction],
        )

        output = tmp_path / "long_name.xlsx"
        export_to_peak([account], output)

        wb = openpyxl.load_workbook(output)
        # Excel max sheet name is 31 chars
        assert len(wb.sheetnames[0]) <= 31
        wb.close()

    def test_each_sheet_has_headers(
        self, sample_transaction, sample_deposit_transaction, tmp_path
    ):
        """Each sheet should have Thai headers."""
        account1 = Account(
            account_number="111-1-11111-1",
            all_transactions=[sample_transaction],
        )
        account2 = Account(
            account_number="222-2-22222-2",
            all_transactions=[sample_deposit_transaction],
        )

        output = tmp_path / "multi_account.xlsx"
        export_to_peak([account1, account2], output)

        wb = openpyxl.load_workbook(output)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            header = (ws.cell(1, 1).value, ws.cell(1, 2).value, ws.cell(1, 3).value)
            assert header == HEADERS
        wb.close()
